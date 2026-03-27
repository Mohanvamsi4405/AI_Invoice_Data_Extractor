"""
Excel handler — all invoice fields as flat columns, one row per invoice.
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .schemas import InvoiceData

logger = logging.getLogger(__name__)

EXCEL_PATH = "data/invoices.xlsx"

# ── Column map: (Excel Header, InvoiceData field | special key) ──────────────
#   special keys: "__sno__", "__filename__", "__extracted_at__", "__items__"
COLUMNS = [
    # Meta
    ("S.No",                    "__sno__"),
    ("File Name",               "__filename__"),
    ("Extracted At",            "__extracted_at__"),

    # ── MANDATORY ─────────────────────────────────────────
    ("Invoice Number",          "invoice_number"),
    ("Invoice Date",            "invoice_date"),
    ("Vendor Name",             "vendor_name"),
    ("Customer Name",           "customer_name"),
    ("Total Amount",            "total_amount"),
    ("Item Details",            "__items__"),          # serialised line items

    # ── DATES ─────────────────────────────────────────────
    ("Due Date",                "due_date"),
    ("Delivery Date",           "delivery_date"),

    # ── VENDOR INFO ───────────────────────────────────────
    ("Vendor Address",          "vendor_address"),
    ("Vendor GSTIN",            "vendor_gstin"),
    ("Vendor PAN",              "vendor_pan"),
    ("Vendor Email",            "vendor_email"),
    ("Vendor Phone",            "vendor_phone"),
    ("Vendor Bank Details",     "vendor_bank_details"),

    # ── CUSTOMER INFO ─────────────────────────────────────
    ("Customer Address",        "customer_address"),
    ("Customer GSTIN",          "customer_gstin"),
    ("Customer Contact",        "customer_contact"),

    # ── FINANCIAL ─────────────────────────────────────────
    ("Subtotal",                "subtotal"),
    ("Discount (Overall)",      "discount_overall"),
    ("Shipping Charges",        "shipping_charges"),
    ("CGST",                    "cgst"),
    ("SGST",                    "sgst"),
    ("IGST",                    "igst"),
    ("Round Off",               "round_off"),
    ("Currency",                "currency"),

    # ── PAYMENT ───────────────────────────────────────────
    ("Payment Method",          "payment_method"),
    ("Transaction ID",          "transaction_id"),
    ("Payment Date",            "payment_date"),
    ("Payment Status",          "payment_status"),

    # ── COMPLIANCE ────────────────────────────────────────
    ("Place of Supply",         "place_of_supply"),
    ("Reverse Charge",          "reverse_charge"),
    ("Notes / Terms",           "notes"),
]

# ── Styling ───────────────────────────────────────────────────────────────────
HDR_BG   = "1E3A5F"
HDR_FG   = "FFFFFF"
ALT_BG   = "EFF4FB"
MAND_BG  = "FFF3CD"   # light amber for mandatory columns header
BORDER_C = "B8C9E1"

MANDATORY_COLS = {
    "Invoice Number","Invoice Date","Vendor Name",
    "Customer Name","Total Amount","Item Details",
}

# column widths (one per COLUMNS entry)
COL_WIDTHS = [
    6, 28, 20,        # meta
    18, 14, 28, 28, 16, 50,  # mandatory
    14, 14,           # dates
    32, 18, 14, 22, 16, 35,  # vendor
    32, 18, 20,       # customer
    14, 16, 16, 12, 12, 12, 10, 10,  # financial
    18, 20, 14, 16,   # payment
    20, 12, 35,       # compliance
]


def _border():
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)


def _init_sheet(ws):
    ws.freeze_panes = "D2"   # freeze S.No, File Name, Extracted At; scroll right
    for i, (label, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=i, value=label)
        bg = MAND_BG if label in MANDATORY_COLS else HDR_BG
        fg = "1E3A5F" if label in MANDATORY_COLS else HDR_FG
        c.font      = Font(name="Calibri", bold=True, color=fg, size=10)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        width = COL_WIDTHS[i - 1] if i <= len(COL_WIDTHS) else 16
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30


def _style_data_row(ws, row_idx: int):
    alt = row_idx % 2 == 0
    for col in range(1, len(COLUMNS) + 1):
        c = ws.cell(row=row_idx, column=col)
        if alt:
            c.fill = PatternFill("solid", fgColor=ALT_BG)
        c.font      = Font(name="Calibri", size=9)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[row_idx].height = 40   # taller rows for wrap


def _serialize_items(items) -> str:
    """Format line items as readable multi-line text for the Excel cell."""
    if not items:
        return ""
    lines = []
    for i, item in enumerate(items, 1):
        parts = []
        if item.description: parts.append(item.description)
        if item.hsn_sac_code: parts.append(f"HSN:{item.hsn_sac_code}")
        if item.quantity:    parts.append(f"Qty:{item.quantity}")
        if item.unit:        parts.append(item.unit)
        if item.unit_price:  parts.append(f"@{item.unit_price}")
        if item.discount:    parts.append(f"Disc:{item.discount}")
        if item.amount:      parts.append(f"={item.amount}")
        lines.append(f"{i}. " + " | ".join(parts))
    return "\n".join(lines)


def _save_workbook(wb, excel_path: str):
    """Save workbook; if file is locked, save to a timestamped fallback path."""
    try:
        wb.save(excel_path)
    except PermissionError:
        from datetime import datetime as _dt
        fallback = excel_path.replace(".xlsx", f"_{_dt.now().strftime('%H%M%S')}.xlsx")
        wb.save(fallback)
        logger.warning(f"Main file locked — saved to fallback: {fallback}")
        raise PermissionError(
            f"invoices.xlsx is open in another program (Excel/IDE). "
            f"Data saved to fallback: {Path(fallback).name}. "
            f"Please close invoices.xlsx and try again."
        )


def save_invoice_to_excel(
    invoice: InvoiceData,
    filename: str,
    excel_path: str = EXCEL_PATH,
) -> int:
    """Append one invoice row. Returns S.No (1-based)."""
    Path(excel_path).parent.mkdir(parents=True, exist_ok=True)

    if Path(excel_path).exists():
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"
        _init_sheet(ws)

    next_row = ws.max_row + 1
    sno      = next_row - 1
    inv_dict = invoice.model_dump()

    for col_idx, (_, field) in enumerate(COLUMNS, 1):
        if field == "__sno__":
            val = sno
        elif field == "__filename__":
            val = filename
        elif field == "__extracted_at__":
            val = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elif field == "__items__":
            val = _serialize_items(invoice.line_items or [])
        else:
            val = inv_dict.get(field) or ""

        ws.cell(row=next_row, column=col_idx, value=val)

    _style_data_row(ws, next_row)
    # Bold S.No
    ws.cell(row=next_row, column=1).font = Font(name="Calibri", bold=True, size=9)

    _save_workbook(wb, excel_path)
    logger.info(f"Excel row {next_row} saved — {filename}")
    return sno


def get_all_invoices(excel_path: str = EXCEL_PATH) -> List[Dict[str, Any]]:
    if not Path(excel_path).exists():
        return []
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    headers = [str(h) if h else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]
